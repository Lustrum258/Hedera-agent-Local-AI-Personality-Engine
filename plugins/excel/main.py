"""
Excel 插件 — 创建、读取、编辑 .xlsx 文件
依赖: openpyxl (pip install openpyxl)

注册工具:
  - excel_create     创建新工作簿/文件
  - excel_read       读取单元格/区域
  - excel_write      写入单元格
  - excel_list       列举工作表和文件信息
  - excel_add_sheet  添加工作表
  - excel_formula    写入公式
"""

import os
import datetime

from hedera.plugin.base import PluginBase

# ─── 依赖检测 ───

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ─── 安全路径 ───

_BASE_DIRS = []


def _init_base_dirs():
    """初始化可写目录列表"""
    global _BASE_DIRS
    if _BASE_DIRS:
        return
    if os.environ.get("HEDERA_DATA_DIR"):
        _BASE_DIRS.append(os.environ["HEDERA_DATA_DIR"])
    # 常见桌面路径
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    if os.path.isdir(desktop):
        _BASE_DIRS.append(desktop)
    _BASE_DIRS.append(os.getcwd())
    _BASE_DIRS.append(os.path.expanduser("~"))


def _safe_path(path: str) -> str:
    """检查路径是否安全（防止目录穿越）"""
    _init_base_dirs()
    path = os.path.normpath(os.path.expanduser(path.strip()))

    # 绝对路径必须落在允许的目录下
    if os.path.isabs(path):
        ok = False
        for base in _BASE_DIRS:
            norm_base = os.path.normpath(base)
            if path.startswith(norm_base + os.sep) or path == norm_base:
                ok = True
                break
        if not ok:
            raise PermissionError(f"不允许的路径: {path}，文件必须在工作目录或桌面下")
        return path

    # 相对路径 → 拼接桌面
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    full = os.path.normpath(os.path.join(desktop, path))
    if full.startswith(os.path.normpath(desktop) + os.sep) or full == os.path.normpath(desktop):
        return full

    raise PermissionError(f"不允许的相对路径: {path}")


def _safe_desktop_path(filename: str) -> str:
    """返回桌面路径"""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    return os.path.join(desktop, filename)


def _xlsx(path: str) -> str:
    """确保有 .xlsx 后缀"""
    if not path.lower().endswith(".xlsx"):
        return path + ".xlsx"
    return path


def _cell_range_to_list(cell_range: str) -> list[tuple]:
    """将 'A1:B3' 这样的区域转成 [(行,列)] 列表"""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    cells = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cells.append((r, c))
    return cells


# ─── 工具函数 ───

def _create_excel(path: str, sheet_name: str = "Sheet1",
                  headers: list = None, data: list[list] = None) -> dict:
    """
    创建新 Excel 文件

    参数:
      path: 文件路径（相对桌面 或 绝对路径）
      sheet_name: 工作表名称（默认 Sheet1）
      headers: 表头行，如 ["姓名", "年龄", "城市"]
      data: 数据行，如 [["张三", 25, "北京"], ["李四", 30, "上海"]]

    返回: {"success": True, "path": "...", "sheets": [...], "cells": N, "download_url": "..."}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl，请运行: pip install openpyxl"}
    try:
        path = _safe_path(_xlsx(path))
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        row_idx = 1
        if headers:
            for col_idx, h in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=h)
                cell.font = Font(bold=True)
            row_idx = 2

        cell_count = 0
        if data:
            for row_data in data:
                for col_idx, val in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=val)
                    cell_count += 1
                row_idx += 1

        wb.save(path)

        result = {
            "success": True,
            "path": path,
            "sheets": wb.sheetnames,
            "rows": (row_idx - 1),
            "cells": (len(headers) if headers else 0) + cell_count,
        }

        # 同时拷贝一份到上传目录，生成下载链接
        try:
            from hedera.core.tools import _uploads_dir as _up
            import shutil, uuid
            if _up:
                dest_dir = os.path.join(_up, "_common")
                os.makedirs(dest_dir, exist_ok=True)
                fname = os.path.basename(path)
                dest = os.path.join(dest_dir, fname)
                if os.path.exists(dest):
                    base, ext = os.path.splitext(fname)
                    n = 1
                    while os.path.exists(os.path.join(dest_dir, f"{base}_{n}{ext}")):
                        n += 1
                    fname = f"{base}_{n}{ext}"
                    dest = os.path.join(dest_dir, fname)
                shutil.copy2(path, dest)
                result["download_url"] = f"/download/_common/{fname}"
        except Exception:
            pass

        return result
    except Exception as e:
        return {"success": False, "error": str(e)}


def _read_excel(path: str, sheet: str = None,
                cell_range: str = None, max_rows: int = 50) -> dict:
    """
    读取 Excel 文件内容

    参数:
      path: 文件路径
      sheet: 工作表名（默认第一个）
      cell_range: 单元格区域，如 "A1:C10"（默认全部）
      max_rows: 最大读取行数

    返回: {"success": True, "data": [[...], ...], "sheets": [...], "sheet": "..."}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl"}
    try:
        path = _safe_path(path)
        if not os.path.exists(path):
            return {"success": False, "error": f"文件不存在: {path}"}

        wb = load_workbook(path, data_only=True)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"工作表不存在: {sheet_name}，可选: {wb.sheetnames}"}

        ws = wb[sheet_name]

        if cell_range:
            from openpyxl.utils import get_column_letter
            # 解析范围
            cells = _cell_range_to_list(cell_range)
            data = []
            last_row = 0
            for r, c in sorted(cells):
                if r > last_row:
                    if data and len(data[-1]) > 0:
                        last_row = r
                        data.append([])
                val = ws.cell(row=r, column=c).value
                if data:
                    data[-1].append(val)
                else:
                    data.append([val])
                last_row = r
            # 按行分组整理
            row_groups = {}
            for r, c in sorted(cells):
                row_groups.setdefault(r, []).append(ws.cell(row=r, column=c).value)
            sorted_rows = sorted(row_groups.keys())
            data = [row_groups[r] for r in sorted_rows[:max_rows]]
        else:
            data = []
            for row in ws.iter_rows(min_row=1, max_row=min(max_rows, ws.max_row or 0),
                                     max_col=min(20, ws.max_column or 0),
                                     values_only=True):
                data.append(list(row))

        wb.close()
        return {
            "success": True,
            "data": data,
            "sheets": wb.sheetnames,
            "sheet": sheet_name,
            "rows": len(data),
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _write_excel(path: str, sheet: str = None,
                 cell: str = None, value=None,
                 data: list[list] = None,
                 start_cell: str = "A1",
                 bold_header: bool = True) -> dict:
    """
    写入 Excel 文件（追加或更新）

    参数:
      path: 文件路径
      sheet: 工作表名（默认第一个）
      cell: 单个单元格，如 "B3"
      value: 单元格值（配合 cell 使用）
      data: 二维数组数据，如 [["A", 1], ["B", 2]]
      start_cell: 数据写入起始位置（默认 A1）
      bold_header: 是否加粗第一行

    返回: {"success": True, "path": "...", "updated": N}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl"}
    try:
        path = _safe_path(_xlsx(path))
        exists = os.path.exists(path)

        if exists:
            wb = load_workbook(path)
        else:
            wb = Workbook()

        sheet_name = sheet or wb.sheetnames[0] if exists else "Sheet1"
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        # 写单格
        if cell and value is not None:
            from openpyxl.utils import coordinate_from_string
            col_letter, row_num = coordinate_from_string(cell)
            ws[cell] = value
            updated = 1

        # 写数据块
        elif data:
            from openpyxl.utils import coordinate_from_string, column_index_from_string
            start_col, start_row = coordinate_from_string(start_cell)
            start_col_idx = column_index_from_string(start_col) if isinstance(start_col, str) else 1
            if isinstance(start_row, str):
                start_row = int(start_row)

            updated = 0
            for i, row_data in enumerate(data):
                for j, val in enumerate(row_data):
                    cell = ws.cell(row=start_row + i, column=start_col_idx + j, value=val)
                    if bold_header and i == 0:
                        cell.font = Font(bold=True)
                    updated += 1
        else:
            return {"success": False, "error": "请提供 cell+value 或 data 参数"}

        wb.save(path)
        wb.close()
        return {
            "success": True,
            "path": path,
            "sheets": wb.sheetnames,
            "updated": updated if 'updated' in dir() else 1,
        }
    except Exception as e:
        return {"success": False, "error": str(e)}


def _list_excel(path: str = None, filename: str = None) -> dict:
    """
    列举 Excel 文件信息

    参数:
      path: 文件路径（查看具体文件）
      filename: 文件名（在桌面上查找）

    返回: 文件列表或文件详情
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl"}

    # 无参 → 列桌面 xlsx
    if not path and not filename:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        try:
            files = sorted([
                f for f in os.listdir(desktop)
                if f.lower().endswith(".xlsx")
            ])
            # 加上文件大小和修改时间
            details = []
            for f in files:
                fp = os.path.join(desktop, f)
                stat = os.stat(fp)
                details.append({
                    "filename": f,
                    "size_kb": stat.st_size / 1024,
                    "modified": datetime.datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                })
            return {"success": True, "files": details, "count": len(details)}
        except Exception as e:
            return {"success": False, "error": str(e)}

    try:
        p = _safe_path(path or filename)
        if not os.path.exists(p):
            return {"success": False, "error": f"文件不存在: {p}"}

        wb = load_workbook(p, read_only=True)
        info = {
            "success": True,
            "path": p,
            "sheets": wb.sheetnames,
            "sheet_count": len(wb.sheetnames),
        }
        wb.close()
        return info
    except Exception as e:
        return {"success": False, "error": str(e)}


def _add_sheet(path: str, sheet_name: str) -> dict:
    """
    添加工作表

    参数:
      path: 文件路径
      sheet_name: 新工作表名称

    返回: {"success": True, "sheets": [...]}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl"}
    try:
        path = _safe_path(_xlsx(path))
        wb = load_workbook(path)
        if sheet_name in wb.sheetnames:
            wb.close()
            return {"success": False, "error": f"工作表已存在: {sheet_name}"}
        ws = wb.create_sheet(sheet_name)
        wb.save(path)
        wb.close()
        return {"success": True, "sheets": wb.sheetnames, "added": sheet_name}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _formula_excel(path: str, sheet: str = None,
                   cell: str = None, formula: str = None) -> dict:
    """
    写入 Excel 公式

    参数:
      path: 文件路径
      sheet: 工作表名
      cell: 单元格，如 "C3"
      formula: 公式（不含 =），如 "SUM(A1:A10)"

    返回: {"success": True}
    """
    if not HAS_OPENPYXL:
        return {"success": False, "error": "缺少 openpyxl"}
    try:
        path = _safe_path(_xlsx(path))
        wb = load_workbook(path)
        sheet_name = sheet or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            return {"success": False, "error": f"工作表不存在: {sheet_name}"}
        ws = wb[sheet_name]
        formula_text = f"={formula}" if not formula.startswith("=") else formula
        ws[cell] = formula_text
        wb.save(path)
        wb.close()
        return {"success": True, "cell": cell, "formula": formula_text}
    except Exception as e:
        return {"success": False, "error": str(e)}


def _find_file(name: str) -> str:
    """在桌面和常见位置查找文件"""
    desktop = os.path.join(os.path.expanduser("~"), "Desktop")
    candidates = [
        os.path.join(desktop, name),
        os.path.join(desktop, _xlsx(name)),
        os.path.join(os.getcwd(), name),
        os.path.join(os.getcwd(), _xlsx(name)),
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return ""


# ─── 插件类 ───

class ExcelPlugin(PluginBase):
    name = "excel"
    description = "Excel 文件创建、读取、编辑，支持单元格读写、公式、工作表管理"
    keywords = ["excel", "表格", "xlsx", "电子表格", "excel文件", "工作簿"]

    def on_load(self, config: dict = None):
        if not HAS_OPENPYXL:
            print("[Excel Plugin] ⚠️  openpyxl 未安装，部分功能不可用。安装: pip install openpyxl")

    def process(self, message: str, context: dict = None) -> str | None:
        msg = message.strip().lower()

        # 检查依赖
        if not HAS_OPENPYXL:
            if any(kw in msg for kw in self.keywords):
                return "openpyxl 没有装，跑不了 Excel。让 Lustrum 跑一下 `pip install openpyxl`。"
            return None

        # 快速列桌面 Excel 文件
        if msg in ("excel", "表格", "excel文件", "有哪些excel"):
            result = _list_excel()
            if result.get("success"):
                files = result.get("files", [])
                if not files:
                    return "桌面上没有 .xlsx 文件。"
                lines = [f"桌面上有 {len(files)} 个 Excel 文件:"]
                for f in files:
                    lines.append(f"  📄 {f['filename']}  ({f['size_kb']:.1f} KB, {f['modified']})")
                return "\n".join(lines)
            return f"查不了: {result.get('error')}"

        return None

    def get_tools(self) -> list[dict]:
        if not HAS_OPENPYXL:
            return []

        return [
            {
                "name": "excel_create",
                "description": "创建新的 Excel 文件。支持设表头、写入数据。文件默认放桌面同时生成下载链接。",
        "result_url": true,
                "fn": _create_excel,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径（相对桌面 或 绝对路径），如 '数据统计'（自动加 .xlsx）"},
                        "sheet_name": {"type": "string", "description": "工作表名称，默认 Sheet1"},
                        "headers": {
                            "type": "array", "items": {"type": "string"},
                            "description": "表头行，如 ['姓名', '年龄', '城市']"
                        },
                        "data": {
                            "type": "array", "items": {"type": "array"},
                            "description": "数据行，如 [['张三', 25, '北京'], ['李四', 30, '上海']]"
                        },
                    },
                    "required": ["path"],
                },
            },
            {
                "name": "excel_read",
                "description": "读取 Excel 文件的内容，支持指定工作表和单元格区域。",
                "fn": _read_excel,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径"},
                        "sheet": {"type": "string", "description": "工作表名（默认第一个）"},
                        "cell_range": {"type": "string", "description": "单元格区域，如 'A1:C10'"},
                        "max_rows": {"type": "integer", "description": "最大读取行数，默认50"},
                    },
                    "required": ["path"],
                },
            },
            {
                "name": "excel_write",
                "description": "写入 Excel 文件。可写单个单元格或写入数据块。文件不存在则自动创建。",
                "fn": _write_excel,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径"},
                        "sheet": {"type": "string", "description": "工作表名（默认第一个）"},
                        "cell": {"type": "string", "description": "单个单元格，如 'B3'"},
                        "value": {"description": "单元格值"},
                        "data": {
                            "type": "array", "items": {"type": "array"},
                            "description": "二维数据，如 [['姓名','分数'], ['张三',95]]"
                        },
                        "start_cell": {"type": "string", "description": "数据写入起始位置，默认 A1"},
                        "bold_header": {"type": "boolean", "description": "是否加粗第一行，默认 true"},
                    },
                    "required": ["path"],
                },
            },
            {
                "name": "excel_list",
                "description": "列出现有 Excel 文件（桌面）或查看指定文件的详细信息（工作表和结构）。",
                "fn": _list_excel,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径（查看指定文件详情，不传则列桌面所有 xlsx）"},
                    },
                },
            },
            {
                "name": "excel_add_sheet",
                "description": "给现有 Excel 文件添加新工作表。",
                "fn": _add_sheet,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径"},
                        "sheet_name": {"type": "string", "description": "新工作表名称"},
                    },
                    "required": ["path", "sheet_name"],
                },
            },
            {
                "name": "excel_formula",
                "description": "在 Excel 单元格中写入公式。公式不带等号，如 'SUM(A1:A10)'。",
                "fn": _formula_excel,
                "parameters": {
                    "type": "object",
                    "properties": {
                        "path": {"type": "string", "description": "文件路径"},
                        "sheet": {"type": "string", "description": "工作表名"},
                        "cell": {"type": "string", "description": "目标单元格，如 'C3'"},
                        "formula": {"type": "string", "description": "公式（不含 =），如 'SUM(A1:A10)'"},
                    },
                    "required": ["path", "cell", "formula"],
                },
            },
        ]

    def get_system_prompt_modifier(self) -> str | None:
        if not HAS_OPENPYXL:
            return None
        return (
            "【Excel 能力】你可以创建、读取、编辑 Excel 文件。"
            "工具：excel_create（新建）、excel_read（读取）、excel_write（写入）、"
            "excel_list（列文件）、excel_add_sheet（加工作表）、excel_formula（公式）。"
            "文件默认放桌面。"
        )
